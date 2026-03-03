import os
import time
import logging
import requests
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
from openpyxl import Workbook
import re  


# === CONFIGURATION ===
FILTER_MODE = 30  # Scrape deals from the last N days
BASE_URL = 'https://www.artemis.bm/deal-directory/'

# === LOGGING ===
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')

# === OUTPUT FOLDER ===
script_dir = os.path.dirname(os.path.abspath(__file__))
output_dir = os.path.join(script_dir, "Artemis_Outputs")
os.makedirs(output_dir, exist_ok=True)

timestamp_now = datetime.now().strftime("%Y%m%d_%H%M%S")
output_path = os.path.join(output_dir, f"artemis_full_data_{timestamp_now}.xlsx")

# === EXCEL SETUP ===
workbook = Workbook()
sheet = workbook.active
sheet.title = "Artemis Deals"

# Define the columns to store
columns = [
    "Issuer", "Cedent / sponsor", "Risks / perils covered", "Size", "Date",
    "Placement / structuring agent/s",
    "Risk modelling / calculation agents etc",
    "Trigger type", "Ratings", "Date of issue",
    "Full Details",
    "Update 1", "Update 2", "Update 3", 
    "Deal Link"

]

# Write header row
for i, col in enumerate(columns, start=1):
    sheet.cell(row=1, column=i).value = col

# === FILTER FUNCTION ===
today = datetime.today()
def get_recent_month_years(days):
    return {(today - timedelta(days=i)).strftime("%b %Y") for i in range(days + 1)}
allowed_months = get_recent_month_years(FILTER_MODE)

# === SCRAPE MAIN TABLE ===
session = requests.Session()
try:
    logging.info("Fetching main Artemis deal directory...")
    soup = BeautifulSoup(session.get(BASE_URL, timeout=15).content, 'html.parser')
    rows = soup.find('table').find_all('tr')[1:]
    logging.info(f"Found {len(rows)} deals in total")
except Exception as e:
    logging.error(f"Failed to load main page: {e}", exc_info=True)
    exit(1)

row_idx = 2
for idx, row in enumerate(rows, start=1):
    try:
        cols = row.find_all('td')
        if len(cols) < 5:
            continue

        issuer_tag = cols[0].find('a', href=True)
        issuer = issuer_tag.text.strip() if issuer_tag else cols[0].text.strip()
        link = issuer_tag['href'] if issuer_tag else None
        cedent, risks, size, date_text = map(lambda x: x.text.strip(), cols[1:5])

        if date_text not in allowed_months:
            continue

        # Build full URL
        full_url = link if link.startswith("http") else f"https://www.artemis.bm{link}"
        logging.info(f"[{idx}] Scraping: {issuer} — {date_text}")

        # Basic deal info
        deal_data = {
            "Issuer": issuer,
            "Cedent / sponsor": cedent,
            "Risks / perils covered": risks,
            "Size": size,
            "Date": date_text,
            "Placement / structuring agent/s": None,
            "Risk modelling / calculation agents etc": None,
            "Trigger type": None,
            "Ratings": None,
            "Date of issue": None,
            "Full Details": None,
            "Update 1": None,
            "Update 2": None,
            "Update 3": None,
            "Deal Link": full_url
        }

        # === SCRAPE DETAIL PAGE ===
        try:
            detail_soup = BeautifulSoup(session.get(full_url, timeout=20).content, 'html.parser')
            info_box = detail_soup.find(id='info-box')

            # 1️⃣ At a glance key–value pairs
            if info_box:
                for li in info_box.find_all('li'):
                    if ':' in li.text:
                        key, val = map(str.strip, li.text.split(':', 1))
                        key_lower = key.lower()
                        for col in deal_data.keys():
                            if col.lower().startswith(key_lower[:10]):  # loose match
                                deal_data[col] = val.strip()
                                break
                        # direct key mapping fallback
                        if key in deal_data:
                            deal_data[key] = val.strip()

            # 2️⃣ Full details paragraph
            full_details_section = None
            for h3 in detail_soup.find_all(['h3', 'h2']):
                if 'Full details' in h3.text:
                    full_details_section = h3.find_next_sibling()
                    break

            
            full_text = ""
            if full_details_section:
                full_text = full_details_section.get_text(separator=' ', strip=True)
            else:
                paragraphs = detail_soup.find_all('p')
                if paragraphs:
                    full_text = ' '.join(p.get_text(strip=True) for p in paragraphs[:5])

            deal_data["Full Details"] = full_text

        # 3️⃣ ✅ Extract "Update 1", "Update 2", "Update 3"
            # if full_text:
            #     # Capture numbered updates, tolerant of spaces or different dash types
            #     pattern = re.compile(
            #         r'Update\s*(\d+)\s*[:\-–]\s*(.*?)(?=(?:Update\s*\d+\s*[:\-–]|$))',
            #         re.IGNORECASE | re.DOTALL
            #     )

            #     matches = pattern.findall(full_text)

            #     # Fill defaults
            #     for i in range(1, 4):
            #         deal_data[f"Update {i}"] = "NA"

            #     for num, text_block in matches:
            #         num = int(num)
            #         if 1 <= num <= 3:
            #             clean_text = re.sub(r'\s+', ' ', text_block).strip()
            #             deal_data[f"Update {num}"] = clean_text

             # 3️⃣ ✅ Extract "Update 1", "Update 2", "Update 3" (non-regex clean version)
            for i in range(1, 4):
                deal_data[f"Update {i}"] = "NA"

            if full_text:
                parts = full_text.split("Update ")
                for part in parts[1:]:
                    num = part[:1]
                    if num.isdigit() and int(num) in [1, 2, 3]:
                        text = part[2:].split("Update ")[0].strip(":-– \n")
                        clean_text = " ".join(text.split())
                        deal_data[f"Update {num}"] = clean_text

        except Exception as e:
            logging.warning(f"Could not fetch full details for {issuer}: {e}")       
       

        # === WRITE TO EXCEL ===
        for col_idx, col_name in enumerate(columns, start=1):
            sheet.cell(row=row_idx, column=col_idx).value = deal_data.get(col_name)
        row_idx += 1

        time.sleep(0.5)  # be polite to website

    except Exception as e:
        logging.error(f"Error processing row {idx}: {e}", exc_info=True)
        continue

# === SAVE OUTPUT ===
try:
    workbook.save(output_path)
    logging.info(f"✅ Excel saved at: {output_path}")
except Exception as e:
    logging.error(f"❌ Error saving Excel: {e}", exc_info=True)
